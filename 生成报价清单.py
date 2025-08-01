#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
极速外贸管理系统报价清单生成器
鸿思特科技有限公司
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_quotation_excel():
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "极速外贸管理系统报价清单"
    
    # 设置样式
    header_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
    title_font = Font(name='微软雅黑', size=16, bold=True, color='2196F3')
    normal_font = Font(name='微软雅黑', size=11)
    bold_font = Font(name='微软雅黑', size=11, bold=True)
    
    header_fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
    sub_header_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    
    # 标题区域
    ws.merge_cells('A1:G1')
    ws['A1'] = '极速外贸管理系统综合版 - 项目报价清单'
    ws['A1'].font = Font(name='微软雅黑', size=18, bold=True, color='2196F3')
    ws['A1'].alignment = center_alignment
    
    # 公司信息
    ws.merge_cells('A2:G2')
    ws['A2'] = '鸿思特科技有限公司 | 联系人：陈志成 | 电话：17600198256 | 邮箱：hst_co@163.com'
    ws['A2'].font = Font(name='微软雅黑', size=12, color='666666')
    ws['A2'].alignment = center_alignment
    
    # 日期
    ws.merge_cells('A3:G3')
    ws['A3'] = f'报价日期：{datetime.now().strftime("%Y年%m月%d日")}'
    ws['A3'].font = normal_font
    ws['A3'].alignment = center_alignment
    
    # 空行
    ws.row_dimensions[4].height = 10
    
    # 表头
    headers = ['序号', '功能模块', '功能描述', '数量', '单价(元)', '小计(元)', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_alignment
        ws.row_dimensions[5].height = 25
    
    # 报价数据
    quotation_data = [
        # 基础模块
        ("基础平台", "系统基础平台搭建", "包含用户管理、权限控制、数据库设计等", 1, 15000, "核心基础"),
        ("控制台模块", "系统首页和导航", "系统介绍、角色概况、业务流程概览", 1, 8000, "标准功能"),
        
        # 核心业务模块
        ("客户管理模块", "客户全生命周期管理", "客户列表、客户公海、查重功能、跟进管理", 1, 25000, "核心模块"),
        ("产品管理模块", "产品档案和供应商管理", "产品信息、供应商管理、同行分析、分类管理", 1, 20000, "核心模块"),
        ("报价管理模块", "智能报价系统", "报价单生成、价格策略、有效期控制、统计分析", 1, 18000, "核心模块"),
        ("订单管理模块", "订单全流程管理", "销售订单、寄样管理、状态跟踪、审批流程", 1, 22000, "核心模块"),
        ("采购管理模块", "采购业务管理", "采购需求、订单管理、供应商协调、成本分析", 1, 20000, "核心模块"),
        ("库存管理模块", "库存全面管控", "库存查询、出入库、盘点、多仓库管理", 1, 18000, "核心模块"),
        ("出运管理模块", "物流出运管理", "出运安排、物流跟踪、运输管理、单据管理", 1, 16000, "核心模块"),
        ("财务管理模块", "财务收付管理", "应收应付、收付款跟踪、利润分析、财务统计", 1, 25000, "核心模块"),
        
        # 分析统计模块
        ("查询统计模块", "数据统计分析", "明细账查询、业务统计、客户分析、业绩统计", 1, 15000, "增值功能"),
        ("报表系统", "智能报表生成", "各类业务报表、图表展示、导出功能", 1, 12000, "增值功能"),
        
        # 系统管理模块
        ("用户权限管理", "多角色权限控制", "用户管理、角色配置、权限分配、部门管理", 1, 10000, "标准功能"),
        ("基础数据管理", "系统基础配置", "币种、支付方式、价格条款、海关编码等", 1, 8000, "标准功能"),
        ("系统维护", "系统运维功能", "数据备份、日志管理、系统监控", 1, 8000, "标准功能"),
        
        # 集成接口
        ("邮件集成", "邮件自动化", "邮件发送、模板管理、自动提醒", 1, 8000, "增值功能"),
        ("短信集成", "短信通知功能", "短信发送、模板配置、消息提醒", 1, 6000, "增值功能"),
        ("微信集成", "微信通知集成", "微信公众号、企业微信消息推送", 1, 10000, "增值功能"),
        ("钉钉集成", "钉钉办公集成", "钉钉消息、审批集成", 1, 8000, "增值功能"),
        
        # 移动端
        ("移动端APP", "手机移动办公", "iOS/Android应用，核心功能移动化", 1, 35000, "高级功能"),
        ("微信小程序", "微信小程序版本", "客户查询、订单跟踪等轻量功能", 1, 20000, "高级功能"),
        
        # 高级功能
        ("数据分析", "BI商业智能", "数据可视化、趋势分析、预测功能", 1, 25000, "高级功能"),
        ("API接口", "第三方系统集成", "标准API接口、数据同步", 1, 15000, "高级功能"),
        
        # 实施服务
        ("系统部署", "系统安装部署", "服务器部署、环境配置、系统调试", 1, 8000, "实施服务"),
        ("数据迁移", "历史数据导入", "旧系统数据迁移、数据清洗", 1, 10000, "实施服务"),
        ("用户培训", "系统使用培训", "操作培训、管理培训（3天）", 1, 6000, "实施服务"),
        ("技术支持", "一年技术支持", "电话支持、远程协助、问题解决", 1, 15000, "售后服务"),
        
        # 定制开发
        ("个性化定制", "特殊需求定制", "根据客户特殊需求定制开发", 1, 20000, "定制服务"),
        ("界面定制", "UI界面定制", "企业VI定制、界面个性化", 1, 8000, "定制服务"),
    ]
    
    # 填充数据
    row = 6
    total_amount = 0
    
    for i, (module, name, desc, qty, price, note) in enumerate(quotation_data, 1):
        subtotal = qty * price
        total_amount += subtotal
        
        # 序号
        ws.cell(row=row, column=1, value=i).font = normal_font
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = center_alignment
        
        # 功能模块
        ws.cell(row=row, column=2, value=module).font = bold_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = left_alignment
        
        # 功能描述
        ws.cell(row=row, column=3, value=f"{name}\n{desc}").font = normal_font
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=3).alignment = left_alignment
        ws.row_dimensions[row].height = 35
        
        # 数量
        ws.cell(row=row, column=4, value=qty).font = normal_font
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).alignment = center_alignment
        
        # 单价
        ws.cell(row=row, column=5, value=f"￥{price:,}").font = normal_font
        ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=5).alignment = center_alignment
        
        # 小计
        ws.cell(row=row, column=6, value=f"￥{subtotal:,}").font = normal_font
        ws.cell(row=row, column=6).border = thin_border
        ws.cell(row=row, column=6).alignment = center_alignment
        
        # 备注
        ws.cell(row=row, column=7, value=note).font = normal_font
        ws.cell(row=row, column=7).border = thin_border
        ws.cell(row=row, column=7).alignment = center_alignment
        
        # 根据备注类型设置背景色
        if note == "核心模块":
            fill_color = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
        elif note == "增值功能":
            fill_color = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
        elif note == "高级功能":
            fill_color = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
        else:
            fill_color = None
            
        if fill_color:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = fill_color
        
        row += 1
    
    # 合计行
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = '合计金额'
    ws[f'A{row}'].font = Font(name='微软雅黑', size=14, bold=True)
    ws[f'A{row}'].border = thin_border
    ws[f'A{row}'].alignment = center_alignment
    ws[f'A{row}'].fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
    
    ws[f'F{row}'] = f"￥{total_amount:,}"
    ws[f'F{row}'].font = Font(name='微软雅黑', size=14, bold=True, color='FF0000')
    ws[f'F{row}'].border = thin_border
    ws[f'F{row}'].alignment = center_alignment
    
    ws[f'G{row}'] = "不含税"
    ws[f'G{row}'].font = normal_font
    ws[f'G{row}'].border = thin_border
    ws[f'G{row}'].alignment = center_alignment
    
    ws.row_dimensions[row].height = 30
    
    # 添加说明信息
    row += 2
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = '报价说明：'
    ws[f'A{row}'].font = Font(name='微软雅黑', size=12, bold=True, color='2196F3')
    
    explanations = [
        "1. 以上报价为人民币含税价格，有效期30天",
        "2. 核心模块为系统必需功能，增值功能和高级功能可根据需求选择",
        "3. 包含完整源代码交付，支持二次开发",
        "4. 免费提供一年技术支持和系统维护",
        "5. 系统支持云部署和本地部署两种方式",
        "6. 开发周期：3-5个月（根据功能选择确定）",
        "7. 付款方式：合同签订后预付50%，系统验收后支付尾款",
        "8. 本报价不包含服务器等硬件设备费用"
    ]
    
    for i, exp in enumerate(explanations):
        row += 1
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = exp
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = left_alignment
    
    # 添加联系信息
    row += 2
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = '如有疑问，请联系我们：'
    ws[f'A{row}'].font = Font(name='微软雅黑', size=12, bold=True, color='2196F3')
    
    row += 1
    contact_info = [
        "公司名称：鸿思特科技有限公司",
        "联 系  人：陈志成",
        "联系电话：17600198256", 
        "电子邮箱：hst_co@163.com",
        "感谢您选择鸿思特科技，我们将为您提供最优质的服务！"
    ]
    
    for info in contact_info:
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = info
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = center_alignment
        row += 1
    
    # 保存文件
    filename = "极速外贸管理系统报价清单.xlsx"
    wb.save(filename)
    print(f"报价清单已生成：{filename}")
    print(f"总金额：￥{total_amount:,}")
    
    return filename

if __name__ == "__main__":
    create_quotation_excel() 